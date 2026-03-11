#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CECOM - Script de Importacion desde Excel
==========================================
Lee los dos archivos Excel de respaldo y genera un archivo SQL para poblar
la base de datos PostgreSQL con datos historicos del 2025.

Requisitos:
    pip install openpyxl

Uso:
    python import_excel.py

Salida:
    cecom_import.sql  ->  psql $DATABASE_URL -f cecom_import.sql
"""

import openpyxl
import re
import sys
import os
from datetime import datetime, date, time as dtime

# Forzar salida UTF-8 en Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# --- RUTAS -------------------------------------------------------------------
BASE = r"C:\Users\ACER\Desktop\Proyectos\excel de datos que pude recopilar"
TIPOLOGIAS_PATH = os.path.join(BASE, "Tipologias CECOM 2025.xlsx")
INCIDENCIAS_PATH = os.path.join(BASE, "Todo_2025_verificar.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cecom_import.sql")

# Nombre alternativo del archivo (con acento)
TIPOLOGIAS_ALT = os.path.join(BASE, "Tipolog\u00edas CECOM 2025.xlsx")

# --- HELPERS SQL -------------------------------------------------------------

def esc(s):
    if s is None or str(s).strip() == '':
        return 'NULL'
    return "'" + str(s).strip().replace("'", "''") + "'"

def esc_coord(v, vmin, vmax):
    """Convierte coordenada detectando automaticamente el divisor correcto."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if s in ('', 'NULL', 'None', 'nan', '0'):
        return 'NULL'
    try:
        f = float(v)
        if f == 0:
            return 'NULL'
        for exp in range(0, 17):
            corrected = f / (10 ** exp)
            if vmin <= corrected <= vmax:
                return "'" + f"{corrected:.8f}" + "'"
        return 'NULL'
    except Exception:
        return 'NULL'

def esc_ts(dt_val):
    if dt_val is None:
        return 'NULL'
    if isinstance(dt_val, datetime):
        return "'" + dt_val.strftime('%Y-%m-%d %H:%M:%S') + "+00'"
    if isinstance(dt_val, date):
        return "'" + dt_val.strftime('%Y-%m-%d') + " 00:00:00+00'"
    s = str(dt_val).strip()
    if s in ('', 'None', 'nan'):
        return 'NULL'
    return "'" + s + "+00'"

def normalize(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip().upper())

def cell_val(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return None
    return v

# --- DETECCION DE ESCALA DE COORDENADAS -------------------------------------

def detect_coord_scale(values, vmin, vmax):
    valid = []
    for v in values:
        if v is not None:
            try:
                f = float(v)
                if f != 0:
                    valid.append(f)
            except Exception:
                pass
    if not valid:
        return 1
    sample = valid[:100]
    for exp in range(0, 16):
        divisor = 10 ** exp
        corrected = [v / divisor for v in sample]
        ok = all(vmin <= c <= vmax for c in corrected)
        if ok:
            return divisor
    return 1

# --- LECTURA TIPOLOGIAS ------------------------------------------------------

def read_tipologias():
    """
    Estructura del Excel (doble header):
      Fila 0: GENERICO | | TIPO | | SUBTIPO | |
      Fila 1: CODIGO | DESCRIPCION | CODIGO | DESCRIPCION | CODIGO | DESCRIPCION | URGENCIA
      Fila 2+: datos

    Columnas:
      0=gen_cod, 1=gen_desc, 2=tip_cod, 3=tip_desc, 4=sub_cod, 5=sub_desc, 6=urgencia
    """
    # Intentar ambas variantes del nombre de archivo
    path = TIPOLOGIAS_PATH
    if not os.path.exists(path):
        path = TIPOLOGIAS_ALT
    if not os.path.exists(path):
        # Buscar en la carpeta
        for f in os.listdir(BASE):
            if 'Tipolog' in f and f.endswith('.xlsx'):
                path = os.path.join(BASE, f)
                break

    print("[*] Leyendo tipologias: " + path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        print("[ERR] Archivo no encontrado: " + path)
        sys.exit(1)

    sheet_name = 'DATA' if 'DATA' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print("  Sheet: " + sheet_name)

    rows = list(ws.iter_rows(values_only=True))
    print("  Total filas: " + str(len(rows)))
    print("  Fila 0: " + str(rows[0]))
    print("  Fila 1: " + str(rows[1]))
    print("  Fila 2: " + str(rows[2]))

    # Indices fijos segun estructura real del Excel
    I_GEN_COD, I_GEN_DESC = 0, 1
    I_TIP_COD, I_TIP_DESC = 2, 3
    I_SUB_COD, I_SUB_DESC = 4, 5
    I_URG = 6

    genericos = {}   # gen_cod_str -> {id, codigo, descripcion}
    tipos = {}       # tip_cod_str -> {id, codigo, descripcion, generico_id}
    subtipos = []

    gen_counter = 1
    tip_counter = 1
    sub_counter = 1
    last_gen_id = None
    last_tip_id = None

    # Datos desde fila 2 (0-indexed), saltando los dos headers
    for row in rows[2:]:
        if len(row) < 5:
            continue

        gen_cod  = cell_val(row, I_GEN_COD)
        gen_desc = cell_val(row, I_GEN_DESC)
        tip_cod  = cell_val(row, I_TIP_COD)
        tip_desc = cell_val(row, I_TIP_DESC)
        sub_cod  = cell_val(row, I_SUB_COD)
        sub_desc = cell_val(row, I_SUB_DESC)
        urg_raw  = cell_val(row, I_URG)

        # GENERICO (herencia hacia abajo: solo cambia cuando hay valor)
        def clean_desc(s):
            if not s:
                return None
            return re.sub(r'[\r\n\t]+', ' ', str(s).strip()).upper()

        if gen_cod is not None:
            key = str(gen_cod).strip()
            if key not in genericos:
                desc = clean_desc(gen_desc) or ('GENERICO ' + key)
                genericos[key] = {
                    'id': gen_counter,
                    'codigo': key.zfill(2),
                    'descripcion': desc,
                }
                gen_counter += 1
            last_gen_id = genericos[key]['id']

        # TIPO (herencia hacia abajo)
        if tip_cod is not None and last_gen_id is not None:
            key = str(tip_cod).strip()
            if key not in tipos:
                desc = clean_desc(tip_desc) or ('TIPO ' + key)
                tipos[key] = {
                    'id': tip_counter,
                    'codigo': key.zfill(4),
                    'descripcion': desc,
                    'generico_id': last_gen_id,
                }
                tip_counter += 1
            last_tip_id = tipos[key]['id']

        # SUBTIPO (cada fila tiene un subtipo)
        if sub_cod is not None and last_tip_id is not None:
            urg = normalize(urg_raw) if urg_raw else 'BAJO'
            desc = clean_desc(sub_desc) or ('SUBTIPO ' + str(sub_cod))
            subtipos.append({
                'id': sub_counter,
                'codigo': str(sub_cod).strip().zfill(6),
                'descripcion': desc,
                'tipo_id': last_tip_id,
                'urgencia': urg,
            })
            sub_counter += 1

    wb.close()
    g_list = list(genericos.values())
    t_list = list(tipos.values())

    print("  [OK] Genericos: " + str(len(g_list)) + "  Tipos: " + str(len(t_list)) + "  Subtipos: " + str(len(subtipos)))
    print("  Genericos encontrados:")
    for g in g_list:
        print("    [" + g['codigo'] + "] " + g['descripcion'][:70])
    print("  Tipos (primeros 5):")
    for t in t_list[:5]:
        print("    [" + t['codigo'] + "] " + t['descripcion'][:60] + " (gen=" + str(t['generico_id']) + ")")
    print("  Subtipos (primeros 3):")
    for s in subtipos[:3]:
        print("    [" + s['codigo'] + "] " + s['descripcion'][:60] + " urg=" + s['urgencia'])

    return g_list, t_list, subtipos

# --- LECTURA INCIDENCIAS -----------------------------------------------------

def read_incidencias(genericos, tipos, subtipos):
    """
    Lee Todo_2025_verificar.xlsx.
    Columnas (0-based):
      0=Codigo, 1=Jurisdiccion, 2=Direccion, 3=SubtipoCod+Desc,
      4=GenericoDesc, 5=Canal, 6=SubtipoDetalle, 7=Fecha,
      8=Anno, 9=Mes, 10=Dia, 11=DiaSemana, 12=Hora,
      13=Descripcion, 14=?, 15=Lat, 16=Lng, 17=Turno,
      18=SerenoNombre, 19=OperadorCod, 20=OperadorNombre
    """
    print("[*] Leyendo incidencias: " + INCIDENCIAS_PATH)
    try:
        wb = openpyxl.load_workbook(INCIDENCIAS_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        print("[ERR] Archivo no encontrado: " + INCIDENCIAS_PATH)
        sys.exit(1)

    ws = wb.active
    # El Excel NO tiene fila de encabezado - todos los registros son datos
    all_rows = list(ws.iter_rows(values_only=True))
    total = len(all_rows)
    print("  Total filas de datos: " + str(total))
    print("  Muestra fila 0: " + str(all_rows[0]))
    print("  Muestra fila 1: " + str(all_rows[1]))

    # Mostrar muestra de coordenadas para verificar
    print("  Coords muestra (col15=lat, col16=lng):")
    for r in all_rows[:3]:
        lat_v = cell_val(r, 15)
        lng_v = cell_val(r, 16)
        print("    lat_raw=" + str(lat_v) + "  lng_raw=" + str(lng_v))
        if lat_v is not None:
            try:
                lat_c = esc_coord(lat_v, -13.0, -11.0)
                lng_c = esc_coord(lng_v, -78.0, -76.0)
                print("    lat_corrected=" + lat_c + "  lng_corrected=" + lng_c)
            except Exception as e:
                print("    Error: " + str(e))

    # Extraer valores unicos para lookups
    # Columnas reales:
    #   col 19 = codigo de operador/camara (p.ej. OPERADOR-CAMARA-2)
    #   col 20 = nombre de persona (sereno o operador, 145 unicos)
    #   col 18 = campo desconocido (ignorar)
    jurisdicciones_set = set()
    serenos_set = set()
    operadores_cod_set = set()
    canales_set = set()
    ignore_vals = {'', 'NONE', 'N/A', '-', 'SIN ASIGNAR', 'NULL', 'NAN'}

    for r in all_rows:
        j = cell_val(r, 1)
        jn = normalize(j) if j else ''
        if jn and jn not in ignore_vals:
            jurisdicciones_set.add(jn)

        # Nombre de persona en col 20 (145 unicos = serenos en CECOM)
        s = cell_val(r, 20)
        sn = normalize(s) if s else ''
        if sn and sn not in ignore_vals:
            serenos_set.add(sn)

        # Codigo de camara/operador en col 19
        op = cell_val(r, 19)
        opn = normalize(op) if op else ''
        if opn and opn not in ignore_vals:
            operadores_cod_set.add(opn)

        c = cell_val(r, 5)
        if c:
            canales_set.add(normalize(c))

    # Filtrar "NULL" de jurisdicciones
    jurisdicciones_set.discard('NULL')

    print("  Jurisdicciones (" + str(len(jurisdicciones_set)) + "): " + str(sorted(jurisdicciones_set)))
    print("  Canales: " + str(sorted(canales_set)))
    print("  Personas/Serenos (col 20): " + str(len(serenos_set)))
    print("  Operadores/Camaras (col 19): " + str(len(operadores_cod_set)))
    print("  Muestra serenos: " + str(sorted(serenos_set)[:5]))

    wb.close()

    # ── Construir lookups de tipologias para el mapeo ─────────────────────
    # TIPO (cols C,D) → tipo_casos en DB: lookup por codigo 4 digitos y descripcion
    tip_by_cod = {}
    tip_by_desc = {}
    for t in tipos:
        tip_by_cod[t['codigo']] = t
        tip_by_cod[t['codigo'].lstrip('0')] = t
        dn = normalize(t['descripcion'])
        tip_by_desc[dn] = t
        words = dn.split()
        if len(words) >= 2:
            tip_by_desc[' '.join(words[:2])] = t

    # SUBTIPO (cols E,F) → sub_tipo_casos en DB: lookup por codigo 6 digitos
    sub_by_cod = {}
    for s in subtipos:
        sub_by_cod[s['codigo']] = s
        sub_by_cod[s['codigo'].lstrip('0')] = s

    # GENERICO → primer tipo_id bajo ese generico (fallback cuando col 4 tiene desc generica)
    gen_by_desc = {}
    gen_to_first_tipo = {}
    for g in genericos:
        dn = normalize(g['descripcion'])
        gen_by_desc[dn] = g['id']
        words = dn.split()
        if len(words) >= 2:
            gen_by_desc[' '.join(words[:2])] = g['id']
        if len(words) >= 3:
            gen_by_desc[' '.join(words[:3])] = g['id']
    for t in tipos:
        gid = t['generico_id']
        if gid not in gen_to_first_tipo:
            gen_to_first_tipo[gid] = t['id']

    return {
        'rows': all_rows,
        'jurisdicciones': sorted(jurisdicciones_set),
        'serenos': sorted(serenos_set),            # nombres col 20
        'operadores': sorted(operadores_cod_set),  # codigos col 19
        'canales': sorted(canales_set),
        'gen_by_desc': gen_by_desc,
        'gen_to_first_tipo': gen_to_first_tipo,
        'tip_by_cod': tip_by_cod,
        'tip_by_desc': tip_by_desc,
        'sub_by_cod': sub_by_cod,
    }

# --- PARSE CODIGO INCIDENCIA -------------------------------------------------

def extract_code(raw):
    """Extrae el primer bloque de 4-6 digitos de un string tipo '0101-DESCRIPCION'."""
    if raw is None:
        return None
    m = re.search(r'\b(\d{4,6})\b', str(raw))
    return m.group(1) if m else None

# --- GENERACION SQL ----------------------------------------------------------

def generate_sql(genericos, tipos, subtipos, inc_data):
    print("[*] Generando SQL -> " + OUTPUT_PATH)

    lines = []
    TS = "CURRENT_TIMESTAMP"

    def w(s):
        lines.append(s)

    w("-- ================================================================")
    w("-- CECOM - Importacion desde Excel")
    w("-- Generado: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    w("-- ================================================================")
    w("")
    w("BEGIN;")
    w("")
    w("-- Limpiar datos existentes (orden correcto para FK)")
    w("TRUNCATE TABLE incidencia_serenos, evidencias, incidencias,")
    w("               sub_tipo_casos, tipo_casos,")
    w("               jurisdicciones, operadores, serenos")
    w("               RESTART IDENTITY CASCADE;")
    w("")

    # ── 1. Jurisdicciones ────────────────────────────────────────────────────
    w("-- ----- 1. JURISDICCIONES --------------------------------------------")
    jur_map = {}
    vals = []
    for i, nombre in enumerate(inc_data['jurisdicciones'], start=1):
        jur_map[nombre] = i
        codigo = "J" + str(i).zfill(2)
        vals.append("  (" + str(i) + ", '" + codigo + "', " + esc(nombre) + ", true, " + TS + ", " + TS + ")")
    if vals:
        w("INSERT INTO jurisdicciones (id, codigo, nombre, habilitado, \"createdAt\", \"updatedAt\") VALUES")
        w(",\n".join(vals) + ";")
        w("SELECT setval('jurisdicciones_id_seq', " + str(len(vals)) + ");")
    w("")

    # ── 2. Unidades (mantener existentes: 1=SERENAZGO, 2=POLICIA, 3=BOMBEROS) ─
    w("-- ----- 2. UNIDADES (mantener seed base: 1=SERENAZGO, 2=POLICIA) ----")
    w("-- Ya insertadas por el seed original. No se modifican.")
    w("")

    # ── 3. TipoCasos desde TIPO (cols C,D del Excel) ────────────────────────
    w("-- ----- 3. TIPO CASOS (desde TIPO de tipologias, cols C-D) -----------")
    w("DELETE FROM sub_tipo_casos;")
    w("DELETE FROM tipo_casos;")
    gen_by_id = {g['id']: g for g in genericos}
    vals = []
    for t in tipos:
        # Derivar unidadId desde la descripcion del GENERICO padre
        gen = gen_by_id.get(t['generico_id'], {})
        gen_desc = (gen.get('descripcion') or '').upper()
        if 'POLICI' in gen_desc:
            unidad_id = 2
        elif 'BOMBERO' in gen_desc or 'INCENDIO' in gen_desc:
            unidad_id = 3
        else:
            unidad_id = 1
        vals.append(
            "  (" + str(t['id']) + ", " + esc(t['codigo']) + ", " + esc(t['descripcion']) + ", " + str(unidad_id) +
            ", true, " + TS + ", " + TS + ")"
        )
    if vals:
        w("INSERT INTO tipo_casos (id, codigo, descripcion, \"unidadId\", habilitado, \"createdAt\", \"updatedAt\") VALUES")
        w(",\n".join(vals) + ";")
        w("SELECT setval('tipo_casos_id_seq', " + str(len(vals)) + ");")
    w("")

    # ── 4. SubTipoCasos desde SUBTIPO (cols E,F del Excel) ──────────────────
    w("-- ----- 4. SUBTIPO CASOS (desde SUBTIPO de tipologias, cols E-F) -----")
    vals = []
    for s in subtipos:
        # tipoCasoId = id del TIPO padre (primeros 4 digitos del codigo)
        vals.append(
            "  (" + str(s['id']) + ", " + esc(s['codigo']) + ", " + esc(s['descripcion']) + ", " + str(s['tipo_id']) +
            ", " + esc(s['urgencia']) + ", true, " + TS + ", " + TS + ")"
        )
    if vals:
        w("INSERT INTO sub_tipo_casos (id, codigo, descripcion, \"tipoCasoId\", urgencia, habilitado, \"createdAt\", \"updatedAt\") VALUES")
        w(",\n".join(vals) + ";")
        w("SELECT setval('sub_tipo_casos_id_seq', " + str(len(vals)) + ");")
    w("")

    # ── 5. Severidades (asegurar que existen) ────────────────────────────────
    w("-- ----- 5. SEVERIDADES -----------------------------------------------")
    w("INSERT INTO severidades (id, descripcion, habilitado, \"createdAt\", \"updatedAt\") VALUES")
    w("  (1, 'BAJA', true, NOW(), NOW()),")
    w("  (2, 'MEDIA', true, NOW(), NOW()),")
    w("  (3, 'ALTA', true, NOW(), NOW()),")
    w("  (4, 'CRITICA', true, NOW(), NOW())")
    w("ON CONFLICT (id) DO NOTHING;")
    w("")

    # ── 6. Estados de incidencia ─────────────────────────────────────────────
    w("-- ----- 6. ESTADOS INCIDENCIA ----------------------------------------")
    w("INSERT INTO estado_incidencias (id, descripcion, habilitado, \"createdAt\", \"updatedAt\") VALUES")
    w("  (1, 'PENDIENTE', true, NOW(), NOW()),")
    w("  (2, 'EN ATENCION', true, NOW(), NOW()),")
    w("  (3, 'ATENDIDA', true, NOW(), NOW()),")
    w("  (4, 'CERRADA', true, NOW(), NOW()),")
    w("  (5, 'CANCELADA', true, NOW(), NOW())")
    w("ON CONFLICT (id) DO NOTHING;")
    w("")

    # ── 7. Serenos ──────────────────────────────────────────────────────────
    w("-- ----- 7. SERENOS ---------------------------------------------------")
    sereno_map = {}
    vals = []
    for i, nombre in enumerate(inc_data['serenos'], start=1):
        sereno_map[nombre] = i
        parts = nombre.split()
        if len(parts) >= 3:
            ap = parts[0]
            am = parts[1]
            nombres = ' '.join(parts[2:])
        elif len(parts) == 2:
            ap = parts[0]
            am = ''
            nombres = parts[1]
        else:
            ap = ''
            am = ''
            nombres = nombre
        vals.append(
            "  (" + str(i) + ", " + esc(nombres) + ", " + esc(ap) + ", " + esc(am) +
            ", 1, true, " + TS + ", " + TS + ")"
        )
    if vals:
        w("INSERT INTO serenos (id, nombres, \"apellidoPaterno\", \"apellidoMaterno\", \"cargoSerenoId\", habilitado, \"createdAt\", \"updatedAt\") VALUES")
        w(",\n".join(vals) + ";")
        w("SELECT setval('serenos_id_seq', " + str(len(vals)) + ");")
    w("")

    # ── 8. Operadores ────────────────────────────────────────────────────────
    w("-- ----- 8. OPERADORES ------------------------------------------------")
    op_map = {}
    op_id_base = 1   # operadores fueron truncados, empezar desde 1
    vals = []
    for i, nombre in enumerate(inc_data['operadores'], start=op_id_base):
        op_map[nombre] = i
        medio = 2 if 'CAM' in nombre or 'CCTV' in nombre else 1
        vals.append(
            "  (" + str(i) + ", " + esc(nombre) + ", " + str(medio) +
            ", true, " + TS + ", " + TS + ")"
        )
    if vals:
        w("INSERT INTO operadores (id, descripcion, \"medioId\", habilitado, \"createdAt\", \"updatedAt\") VALUES")
        w(",\n".join(vals) + ";")
        w("SELECT setval('operadores_id_seq', " + str(op_id_base + len(vals) - 1) + ");")
    w("")

    # ── 9. Incidencias (batches de 500) ──────────────────────────────────────
    w("-- ----- 9. INCIDENCIAS (historico 2025) ------------------------------")
    w("")

    rows_data = inc_data['rows']
    gen_by_desc = inc_data['gen_by_desc']
    gen_to_first_tipo = inc_data['gen_to_first_tipo']
    tip_by_cod = inc_data['tip_by_cod']
    sub_by_cod = inc_data['sub_by_cod']

    CANAL_UNIDAD = {'POLICIAL': 2, 'SERENO': 1}
    CANAL_MEDIO  = {'POLICIAL': 2, 'SERENO': 1}
    ESTADO_ATENDIDA = 3   # historico = ATENDIDA

    def urg_to_sev(u):
        u = normalize(u)
        if 'ALT' in u:
            return 3
        if 'MED' in u:
            return 2
        return 1

    # Mapa subtipo (6-digit code) -> severidad e id del TIPO padre
    sub_sev = {s['codigo']: urg_to_sev(s['urgencia']) for s in subtipos}
    sub_tipo_id_map = {s['codigo']: s['tipo_id'] for s in subtipos}

    cols = (
        '"codigoIncidencia"', '"unidadId"', '"tipoCasoId"', '"subTipoCasoId"',
        'direccion', 'latitud', 'longitud', 'descripcion',
        '"registradoEn"', '"ocurridoEn"',
        '"jurisdiccionId"', '"situacionId"', '"medioId"', '"operadorId"',
        '"severidadId"', '"createdAt"', '"updatedAt"'
    )

    # Deduplicar por codigoIncidencia (mantener primera ocurrencia)
    seen_codigos = set()
    unique_rows = []
    dupes = 0
    for r in rows_data:
        cod = cell_val(r, 0)
        if cod is None or str(cod).strip() == '':
            continue
        cod_str = str(cod).strip()
        if cod_str in seen_codigos:
            dupes += 1
            continue
        seen_codigos.add(cod_str)
        unique_rows.append(r)
    print("  Filas unicas: " + str(len(unique_rows)) + "  Duplicados eliminados: " + str(dupes))
    rows_data = unique_rows

    batch_size = 500
    total = len(rows_data)
    batches = (total + batch_size - 1) // batch_size
    inserted = 0
    skipped = 0

    for batch_num in range(batches):
        start_i = batch_num * batch_size
        end_i = min(start_i + batch_size, total)
        batch = rows_data[start_i:end_i]

        val_rows = []
        for row in batch:
            codigo    = cell_val(row, 0)
            jur_raw   = cell_val(row, 1)
            direccion = cell_val(row, 2)
            sub_raw      = cell_val(row, 3)   # codigo TIPO (4 dig): "S - 0204-..."
            gen_raw      = cell_val(row, 4)   # descripcion GENERICO: "Alteracion de Orden Publico"
            subtipo_raw  = cell_val(row, 6)   # codigo SUBTIPO (6 dig): "020401-PRESUNTO..."
            canal_raw    = cell_val(row, 5)
            fecha_raw = cell_val(row, 7)
            hora_raw  = cell_val(row, 12)
            desc_raw  = cell_val(row, 13)
            lat_raw   = cell_val(row, 15)
            lng_raw   = cell_val(row, 16)
            op_raw    = cell_val(row, 19)   # codigo camara/operador
            ser_raw   = cell_val(row, 20)   # nombre persona/sereno

            if codigo is None or str(codigo).strip() == '':
                skipped += 1
                continue

            gen_norm = normalize(gen_raw)  # col 4 = descripcion del GENERICO
            tipo_id = None
            sub_tipo_id = None
            sev_id = None

            # 1. Col 3: codigo de TIPO (4 digitos, ej: "S - 0204-...") -> tipoCasoId
            tipo_code = extract_code(sub_raw)
            if tipo_code:
                t_entry = tip_by_cod.get(tipo_code.zfill(4)) or tip_by_cod.get(tipo_code)
                if t_entry:
                    tipo_id = t_entry['id']

            # 2. Col 6: codigo de SUBTIPO (6 digitos, ej: "020401-PRESUNTO...") -> subTipoCasoId
            subtipo_code = extract_code(subtipo_raw)
            if subtipo_code:
                s_entry = sub_by_cod.get(subtipo_code.zfill(6)) or sub_by_cod.get(subtipo_code)
                if s_entry:
                    sub_tipo_id = s_entry['id']
                    sev_id = urg_to_sev(s_entry['urgencia'])
                    if not tipo_id:
                        tipo_id = s_entry['tipo_id']  # tipo padre del subtipo

            # 3. Fallback: usar col 4 (descripcion generica) para encontrar tipo_id
            if not tipo_id:
                for key in [gen_norm, ' '.join(gen_norm.split()[:3]), ' '.join(gen_norm.split()[:2])]:
                    if key in gen_by_desc:
                        gid = gen_by_desc[key]
                        tipo_id = gen_to_first_tipo.get(gid)
                        break

            canal_norm = normalize(canal_raw) if canal_raw else ''
            unidad_id = CANAL_UNIDAD.get(canal_norm, 1)
            medio_id = CANAL_MEDIO.get(canal_norm, 1)
            jur_id = jur_map.get(normalize(jur_raw)) if jur_raw else None
            op_norm = normalize(op_raw) if op_raw else None
            op_id = op_map.get(op_norm) if op_norm else None
            # Sereno no se inserta en incidencias directamente (va en incidencia_serenos)
            # Por ahora omitimos la asignacion de sereno a incidencia

            # Fecha/hora
            ocurrido = None
            if isinstance(fecha_raw, datetime):
                ocurrido = fecha_raw
            elif isinstance(fecha_raw, date):
                if isinstance(hora_raw, dtime):
                    ocurrido = datetime(fecha_raw.year, fecha_raw.month, fecha_raw.day,
                                       hora_raw.hour, hora_raw.minute, hora_raw.second)
                else:
                    ocurrido = datetime(fecha_raw.year, fecha_raw.month, fecha_raw.day)

            lat_sql = esc_coord(lat_raw, -13.0, -11.0)
            lng_sql = esc_coord(lng_raw, -78.0, -76.0)
            desc_sql = esc(str(desc_raw)[:2000] if desc_raw else None)
            dir_sql = esc(str(direccion)[:500] if direccion else None)

            val_rows.append(
                "  (" +
                esc(str(codigo)) + ", " +
                str(unidad_id) + ", " +
                (str(tipo_id) if tipo_id else 'NULL') + ", " +
                (str(sub_tipo_id) if sub_tipo_id else 'NULL') + ", " +
                dir_sql + ", " +
                lat_sql + ", " +
                lng_sql + ", " +
                desc_sql + ", " +
                esc_ts(ocurrido) + ", " +
                esc_ts(ocurrido) + ", " +
                (str(jur_id) if jur_id else 'NULL') + ", " +
                str(ESTADO_ATENDIDA) + ", " +
                str(medio_id) + ", " +
                (str(op_id) if op_id else 'NULL') + ", " +
                (str(sev_id) if sev_id else 'NULL') + ", " +
                "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP" +
                ")"
            )
            inserted += 1

        if val_rows:
            w("-- Batch " + str(batch_num+1) + "/" + str(batches) +
              " (filas " + str(start_i+1) + "-" + str(end_i) + ")")
            w("INSERT INTO incidencias (" + ", ".join(cols) + ") VALUES")
            w(",\n".join(val_rows) + ";")
            w("")

        if (batch_num + 1) % 20 == 0 or batch_num == batches - 1:
            pct = (batch_num + 1) / batches * 100
            print("  [...] " + str(round(pct)) + "% (" + str(inserted) + " filas)", end='\r')

    print("\n  [OK] Filas preparadas: " + str(inserted) + "  Saltadas: " + str(skipped))

    w("-- ----- Reset sequences ----------------------------------------------")
    w("SELECT setval('incidencias_id_seq', (SELECT COALESCE(MAX(id), 1) FROM incidencias));")
    w("")
    w("COMMIT;")
    w("")
    w("-- [OK] Importacion completa: " + str(inserted) + " incidencias")

    sql_content = '\n'.join(lines)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(sql_content)

    size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
    print("  Archivo: " + OUTPUT_PATH + " (" + str(round(size_mb, 1)) + " MB)")

# --- MAIN --------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  CECOM - Importacion desde Excel")
    print("=" * 60)

    genericos, tipos, subtipos = read_tipologias()
    inc_data = read_incidencias(genericos, tipos, subtipos)
    generate_sql(genericos, tipos, subtipos, inc_data)

    print("")
    print("=" * 60)
    print("[OK] COMPLETADO")
    print("=" * 60)
    print("")
    print("Para importar a PostgreSQL:")
    print("  psql $DATABASE_URL -f " + OUTPUT_PATH)
    print("")
    print("En Windows PowerShell:")
    print("  cd C:/Users/ACER/Desktop/Proyectos/cecom-backend")
    print("  $url = (gc .env | sls DATABASE_URL).ToString().Split('=',2)[1]")
    print("  psql $url -f scripts/cecom_import.sql")

if __name__ == '__main__':
    main()
